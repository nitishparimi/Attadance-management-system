import cv2
from pyzbar import pyzbar
import openpyxl
import datetime
import os

# Global variable to store the Excel workbook
wb = None
excel_file = None

def scan_barcode_from_image(image_path):
    image = cv2.imread(image_path)

    if image is None:
        print("Error loading the image. Please check the file path.")
        return []

    decoded_objects = pyzbar.decode(image)

    if decoded_objects:
        barcode_data_list = []
        for obj in decoded_objects:
            barcode_data = obj.data.decode('utf-8')
            barcode_type = obj.type
            print(f"Detected barcode: {barcode_data}")
            print(f"Barcode type: {barcode_type}\n")
            barcode_data_list.append(barcode_data)

        return barcode_data_list
    else:
        print("No barcodes found in the image.")
        return []

def save_to_excel(barcode_data_list):
    global wb, excel_file
    if wb is None:
        create_new_excel_file()

    ws = wb.active
    start_row = ws.max_row + 1

    for i, barcode_data in enumerate(barcode_data_list, start=start_row):
        ws.cell(row=i, column=1, value=barcode_data)
        print(f"Barcode data saved: {barcode_data}")

    # Save the workbook after writing the data
    wb.save(excel_file)
    print("Excel file updated with new data.")

def create_new_excel_file():
    global wb, excel_file
    current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    excel_file = f"barcodes_{current_time}.xlsx"
    wb = openpyxl.Workbook()
    wb.save(excel_file)
    print(f"New Excel file created: {excel_file}")

def upload_images():
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        print("Please install the 'tkinter' package to use the file dialog.")
        return

    create_new_excel_file()

    while True:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        file_path = filedialog.askopenfilename(title="Select an image file",
                                               filetypes=[("Image files", "*.jpg *.jpeg *.png")])
        if file_path:
            barcode_data_list = scan_barcode_from_image(file_path)
            if barcode_data_list:
                save_to_excel(barcode_data_list)
        else:
            print("No image selected. Exiting the program.")
            break

if __name__ == "__main__":
    upload_images()
